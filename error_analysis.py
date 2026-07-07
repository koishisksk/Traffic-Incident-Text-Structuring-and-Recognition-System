from __future__ import annotations

import argparse
from collections import Counter
from pathlib import Path
from typing import Any

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError as exc:  # pragma: no cover - environment guard
    raise SystemExit("缺少 openpyxl，无法生成 error_report.xlsx") from exc

from evaluator import (
    DEFAULT_SAMPLES_PATH,
    RAW_TEXT_SHEET,
    build_field_stats,
    read_sheet_records,
    read_standard_records,
    read_system_records,
    values_equal,
)


ERROR_COLUMNS = [
    "id",
    "source_text",
    "field_name",
    "standard_value",
    "predicted_value",
    "error_type",
]

SUGGESTION_MAP = {
    "event_type": "补充事件类型关键词和优先级规则，重点处理同一句中多类关键词共现时的判定顺序。",
    "event_subtype": "扩展子类型词典，增加追尾、刮碰、车辆故障、占道施工、信号灯故障等更细粒度触发词。",
    "event_time_text": "增加早高峰、晚高峰、近日、目前、上午/下午口语表达等时间规则。",
    "location": "优化地点边界识别，优先抽取路口、方向、入口、桥面、隧道内、学校门口等完整位置短语。",
    "road_name": "补充道路后缀和交叉路口规则，支持“X路与Y街”“高架”“桥”“隧道”等多道路表达。",
    "vehicle_count": "增强车辆数量归一化，识别“两车”“多车”“一辆货车与小客车”等隐含数量。",
    "casualties": "补充伤亡同义表达，如倒地、送医、轻微受伤、无人员受伤、暂无伤亡等。",
    "road_impact": "扩展道路影响词典，区分占道、停驶、车辆避让、通行受阻、影响通行等表达。",
    "congestion_level": "细化拥堵等级规则，将短时拥堵、排队明显、通行缓慢、严重拥堵映射为统一等级。",
    "occupied_lane": "补充车道抽取规则，识别最左侧、右侧车道、应急车道、直行车道、非机动车道等。",
    "disposal_measure": "扩展处置措施规则，识别交警到场、清障、抢修、分流、绕行、急救到场等表达。",
    "event_status": "增加状态判定规则，识别正在处置、已恢复、已解除、持续施工、故障排除等状态。",
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="交通事件结构化识别错误分析")
    parser.add_argument(
        "--samples",
        default=str(DEFAULT_SAMPLES_PATH),
        help="样本 Excel 路径，默认 outputs/traffic_samples/samples.xlsx",
    )
    parser.add_argument(
        "--system-sheet",
        help="系统输出所在工作表名；不填写时自动查找，找不到则调用当前系统生成输出",
    )
    parser.add_argument(
        "--output",
        default="error_report.xlsx",
        help="错误报告输出路径，默认 error_report.xlsx",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    report = analyze_errors(
        samples_path=Path(args.samples),
        output_path=Path(args.output),
        system_sheet=args.system_sheet,
    )
    print(f"错误报告已生成：{report['output_path']}")
    print(f"系统输出来源：{report['system_output_source']}")
    print(f"错误明细数：{report['error_count']}")
    print(f"事件类型错误数：{report['event_type_error_count']}")
    print(f"最容易出错字段：{report['top_error_fields']}")


def analyze_errors(
    samples_path: Path,
    output_path: Path,
    system_sheet: str | None = None,
) -> dict[str, Any]:
    if not samples_path.exists():
        raise FileNotFoundError(f"找不到样本文件：{samples_path}")

    workbook = openpyxl.load_workbook(samples_path, data_only=True)
    raw_records = read_sheet_records(workbook, RAW_TEXT_SHEET)
    standard_records = read_standard_records(workbook, raw_records)
    system_records, system_source, json_valid_map = read_system_records(workbook, raw_records, system_sheet)

    raw_by_id = {str(record.get("id") or index + 1): record for index, record in enumerate(raw_records)}
    standard_by_id = {record["id"]: record for record in standard_records}
    system_by_id = {record["id"]: record for record in system_records}
    field_names = ["event_type", *build_field_stats(standard_records).keys()]

    error_rows: list[dict[str, Any]] = []
    field_error_counter: Counter[str] = Counter()
    event_type_error_count = 0

    for sample_id in standard_by_id:
        standard = standard_by_id[sample_id]
        predicted = system_by_id.get(sample_id, {})
        source_text = raw_by_id.get(sample_id, {}).get("source_text")

        if not json_valid_map.get(sample_id, False):
            error_rows.append(
                build_error_row(
                    sample_id=sample_id,
                    source_text=source_text,
                    field_name="__json_format__",
                    standard_value="标准 JSON 字段完整",
                    predicted_value="系统输出 JSON 格式错误或字段不完整",
                    error_type="json_format_error",
                )
            )
            field_error_counter["__json_format__"] += 1

        for field_name in field_names:
            standard_value = standard.get(field_name)
            predicted_value = predicted.get(field_name)
            if values_equal(predicted_value, standard_value):
                continue

            error_type = classify_error(field_name, standard_value, predicted_value)
            if field_name == "event_type":
                event_type_error_count += 1
            field_error_counter[field_name] += 1
            error_rows.append(
                build_error_row(
                    sample_id=sample_id,
                    source_text=source_text,
                    field_name=field_name,
                    standard_value=standard_value,
                    predicted_value=predicted_value,
                    error_type=error_type,
                )
            )

    field_stats = build_error_stats(field_error_counter, len(standard_records))
    suggestions = build_suggestions(field_stats)
    write_error_report(output_path, error_rows, field_stats, suggestions, system_source)

    return {
        "output_path": str(output_path),
        "system_output_source": system_source,
        "sample_count": len(standard_records),
        "event_type_error_count": event_type_error_count,
        "error_count": len(error_rows),
        "top_error_fields": [item["field_name"] for item in field_stats[:5]],
    }


def build_error_row(
    sample_id: str,
    source_text: Any,
    field_name: str,
    standard_value: Any,
    predicted_value: Any,
    error_type: str,
) -> dict[str, Any]:
    return {
        "id": sample_id,
        "source_text": stringify(source_text),
        "field_name": field_name,
        "standard_value": stringify(standard_value),
        "predicted_value": stringify(predicted_value),
        "error_type": error_type,
    }


def classify_error(field_name: str, standard_value: Any, predicted_value: Any) -> str:
    if field_name == "event_type":
        return "event_type_mismatch"
    if is_empty(standard_value) and not is_empty(predicted_value):
        return "false_positive"
    if not is_empty(standard_value) and is_empty(predicted_value):
        return "missing_extraction"
    return "value_mismatch"


def build_error_stats(counter: Counter[str], sample_count: int) -> list[dict[str, Any]]:
    stats = []
    for field_name, error_count in counter.most_common():
        stats.append(
            {
                "field_name": field_name,
                "error_count": error_count,
                "sample_count": sample_count,
                "error_rate": round(error_count / sample_count, 4) if sample_count else 0.0,
            }
        )
    return stats


def build_suggestions(field_stats: list[dict[str, Any]]) -> list[dict[str, Any]]:
    suggestions = []
    for index, stats in enumerate(field_stats, start=1):
        field_name = stats["field_name"]
        suggestion = SUGGESTION_MAP.get(field_name, "检查该字段的规则覆盖范围，补充关键词、正则边界和空值处理逻辑。")
        suggestions.append(
            {
                "rank": index,
                "field_name": field_name,
                "error_count": stats["error_count"],
                "error_rate": stats["error_rate"],
                "suggestion": suggestion,
            }
        )
    return suggestions


def write_error_report(
    output_path: Path,
    error_rows: list[dict[str, Any]],
    field_stats: list[dict[str, Any]],
    suggestions: list[dict[str, Any]],
    system_source: str,
) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    report_workbook = openpyxl.Workbook()

    detail_sheet = report_workbook.active
    detail_sheet.title = "Error_Details"
    append_table(detail_sheet, ERROR_COLUMNS, error_rows)

    stats_sheet = report_workbook.create_sheet("Field_Stats")
    append_table(stats_sheet, ["field_name", "error_count", "sample_count", "error_rate"], field_stats)

    suggestion_sheet = report_workbook.create_sheet("Suggestions")
    append_table(suggestion_sheet, ["rank", "field_name", "error_count", "error_rate", "suggestion"], suggestions)

    summary_sheet = report_workbook.create_sheet("Summary")
    summary_rows = [
        {"metric": "system_output_source", "value": system_source},
        {"metric": "total_error_rows", "value": len(error_rows)},
        {"metric": "event_type_error_rows", "value": next((item["error_count"] for item in field_stats if item["field_name"] == "event_type"), 0)},
        {"metric": "top_error_field", "value": field_stats[0]["field_name"] if field_stats else None},
    ]
    append_table(summary_sheet, ["metric", "value"], summary_rows)

    for sheet in report_workbook.worksheets:
        style_sheet(sheet)

    report_workbook.save(output_path)


def append_table(sheet: Any, headers: list[str], rows: list[dict[str, Any]]) -> None:
    sheet.append(headers)
    for row in rows:
        sheet.append([row.get(header) for header in headers])


def style_sheet(sheet: Any) -> None:
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    for column_cells in sheet.columns:
        column_letter = get_column_letter(column_cells[0].column)
        max_length = max(len(stringify(cell.value) or "") for cell in column_cells)
        sheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 60)

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions


def stringify(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, list):
        return "；".join(str(item) for item in value)
    if isinstance(value, dict):
        return str(value)
    return str(value)


def is_empty(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, str):
        return value.strip().lower() in {"", "null", "none", "nan"}
    if isinstance(value, (list, tuple, set, dict)):
        return len(value) == 0
    return False


if __name__ == "__main__":
    main()
