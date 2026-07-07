from __future__ import annotations

import argparse
import json
import re
from pathlib import Path
from typing import Any

try:
    import openpyxl
except ImportError as exc:  # pragma: no cover - environment guard
    raise SystemExit("缺少 openpyxl，无法读取 samples.xlsx") from exc

from src.pipeline import TrafficEventPipeline
from src.schema import OUTPUT_FIELDS


DEFAULT_SAMPLES_PATH = Path("outputs") / "traffic_samples" / "samples.xlsx"
STANDARD_SHEET = "Standard_JSON"
RAW_TEXT_SHEET = "Raw_Texts"
SYSTEM_SHEET_CANDIDATES = [
    "System_Output",
    "System_Outputs",
    "System_JSON",
    "Predictions",
    "Prediction",
    "Output",
    "Outputs",
    "系统输出",
]
JSON_COLUMN_CANDIDATES = [
    "system_output",
    "system_json",
    "output_json",
    "prediction_json",
    "json",
    "系统输出",
]
FIELD_ALIASES = {
    "event_time": "event_time_text",
    "TIME": "event_time_text",
    "LOCATION": "location",
    "ROAD": "road_name",
    "VEHICLE": "vehicles_involved",
    "CASUALTY": "casualties",
    "ROAD_IMPACT": "road_impact",
    "CONGESTION": "congestion_level",
    "DISPOSAL": "disposal_measure",
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="交通事件结构化识别评估器")
    parser.add_argument(
        "--samples",
        default=str(DEFAULT_SAMPLES_PATH),
        help="样本 Excel 路径，默认 outputs/traffic_samples/samples.xlsx",
    )
    parser.add_argument(
        "--system-sheet",
        help="系统输出所在工作表名；不填写时自动查找，找不到则调用当前系统生成输出",
    )
    parser.add_argument(
        "--save-report",
        help="可选：将评估报告保存为 JSON 文件",
    )
    parser.add_argument(
        "--show-errors",
        action="store_true",
        help="输出字段不匹配明细",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    workbook_path = Path(args.samples)
    report = evaluate_workbook(workbook_path, system_sheet=args.system_sheet, show_errors=args.show_errors)
    report_text = json.dumps(report, ensure_ascii=False, indent=2)
    print(report_text)

    if args.save_report:
        output_path = Path(args.save_report)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        output_path.write_text(report_text, encoding="utf-8")


def evaluate_workbook(
    workbook_path: Path,
    system_sheet: str | None = None,
    show_errors: bool = False,
) -> dict[str, Any]:
    if not workbook_path.exists():
        raise FileNotFoundError(f"找不到样本文件：{workbook_path}")

    workbook = openpyxl.load_workbook(workbook_path, data_only=True)
    raw_records = read_sheet_records(workbook, RAW_TEXT_SHEET)
    standard_records = read_standard_records(workbook, raw_records)
    system_records, system_source, json_valid_map = read_system_records(workbook, raw_records, system_sheet)

    standard_by_id = {record["id"]: record for record in standard_records}
    system_by_id = {record["id"]: record for record in system_records}
    sample_ids = [record["id"] for record in standard_records]

    type_total = len(sample_ids)
    type_correct = 0
    json_correct = 0
    field_stats = build_field_stats(standard_records)
    errors: list[dict[str, Any]] = []

    for sample_id in sample_ids:
        standard = standard_by_id[sample_id]
        system = system_by_id.get(sample_id, {})

        if values_equal(system.get("event_type"), standard.get("event_type")):
            type_correct += 1

        if json_valid_map.get(sample_id, False):
            json_correct += 1

        for field in field_stats:
            expected = standard.get(field)
            predicted = system.get(field)
            is_correct = values_equal(predicted, expected)
            field_stats[field]["total"] += 1
            if is_correct:
                field_stats[field]["correct"] += 1
            elif show_errors:
                errors.append(
                    {
                        "id": sample_id,
                        "field": field,
                        "expected": expected,
                        "predicted": predicted,
                    }
                )

    field_total = sum(item["total"] for item in field_stats.values())
    field_correct = sum(item["correct"] for item in field_stats.values())
    field_detail = {
        field: {
            "accuracy": safe_rate(stats["correct"], stats["total"]),
            "correct": stats["correct"],
            "total": stats["total"],
        }
        for field, stats in field_stats.items()
    }

    report: dict[str, Any] = {
        "samples_file": str(workbook_path),
        "system_output_source": system_source,
        "sample_count": len(sample_ids),
        "event_type_accuracy": safe_rate(type_correct, type_total),
        "field_extraction_accuracy": safe_rate(field_correct, field_total),
        "json_format_accuracy": safe_rate(json_correct, len(sample_ids)),
        "counts": {
            "event_type_correct": type_correct,
            "event_type_total": type_total,
            "field_correct": field_correct,
            "field_total": field_total,
            "json_format_correct": json_correct,
            "json_format_total": len(sample_ids),
        },
        "field_detail": field_detail,
    }
    if show_errors:
        report["errors"] = errors
    return report


def read_sheet_records(workbook: Any, sheet_name: str) -> list[dict[str, Any]]:
    if sheet_name not in workbook.sheetnames:
        return []

    sheet = workbook[sheet_name]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [normalize_header(value) for value in rows[0]]
    records = []
    for row in rows[1:]:
        if not any(value is not None for value in row):
            continue
        record = {}
        for index, header in enumerate(headers):
            if header:
                record[header] = normalize_cell(row[index] if index < len(row) else None)
        records.append(record)
    return records


def read_standard_records(workbook: Any, raw_records: list[dict[str, Any]]) -> list[dict[str, Any]]:
    standard_records = read_sheet_records(workbook, STANDARD_SHEET)
    if not standard_records:
        raise ValueError(f"工作簿缺少标准答案工作表：{STANDARD_SHEET}")

    raw_ids = [str(record.get("id") or index + 1) for index, record in enumerate(raw_records)]
    normalized_records = []
    for index, record in enumerate(standard_records):
        normalized = normalize_record_fields(record)
        normalized["id"] = str(record.get("id") or (raw_ids[index] if index < len(raw_ids) else index + 1))
        normalized_records.append(normalized)
    return normalized_records


def read_system_records(
    workbook: Any,
    raw_records: list[dict[str, Any]],
    system_sheet: str | None,
) -> tuple[list[dict[str, Any]], str, dict[str, bool]]:
    sheet_name = system_sheet or find_system_sheet(workbook)
    if sheet_name:
        records = read_sheet_records(workbook, sheet_name)
        return parse_system_sheet_records(records, raw_records, sheet_name)

    pipeline = TrafficEventPipeline()
    generated_records = []
    json_valid_map = {}
    for index, raw in enumerate(raw_records):
        sample_id = str(raw.get("id") or index + 1)
        text = str(raw.get("source_text") or "")
        try:
            output = dict(pipeline.parse(text))
            output["id"] = sample_id
            generated_records.append(output)
            json_valid_map[sample_id] = is_valid_output_json(output)
        except Exception:
            generated_records.append({"id": sample_id})
            json_valid_map[sample_id] = False
    return generated_records, "generated_by_current_pipeline", json_valid_map


def parse_system_sheet_records(
    records: list[dict[str, Any]],
    raw_records: list[dict[str, Any]],
    sheet_name: str,
) -> tuple[list[dict[str, Any]], str, dict[str, bool]]:
    raw_ids = [str(record.get("id") or index + 1) for index, record in enumerate(raw_records)]
    parsed_records = []
    json_valid_map = {}

    for index, record in enumerate(records):
        sample_id = str(record.get("id") or (raw_ids[index] if index < len(raw_ids) else index + 1))
        json_text = first_present(record, JSON_COLUMN_CANDIDATES)

        if json_text is not None:
            parsed, is_valid = parse_json_output(json_text)
            parsed["id"] = sample_id
            parsed_records.append(normalize_record_fields(parsed))
            json_valid_map[sample_id] = is_valid and is_valid_output_json(parsed)
            continue

        normalized = normalize_record_fields(record)
        normalized["id"] = sample_id
        parsed_records.append(normalized)
        json_valid_map[sample_id] = is_valid_output_json(normalized)

    return parsed_records, f"sheet:{sheet_name}", json_valid_map


def find_system_sheet(workbook: Any) -> str | None:
    for candidate in SYSTEM_SHEET_CANDIDATES:
        if candidate in workbook.sheetnames:
            return candidate
    return None


def build_field_stats(standard_records: list[dict[str, Any]]) -> dict[str, dict[str, int]]:
    fields = []
    for record in standard_records:
        for field in record:
            if field not in {"id", "source_text", "event_type", "event_type_code"} and field not in fields:
                fields.append(field)
    return {field: {"correct": 0, "total": 0} for field in fields}


def parse_json_output(value: Any) -> tuple[dict[str, Any], bool]:
    if isinstance(value, dict):
        return value, True
    if value is None:
        return {}, False

    text = str(value).strip()
    if not text:
        return {}, False
    try:
        parsed = json.loads(text)
    except json.JSONDecodeError:
        return {}, False
    return (parsed, True) if isinstance(parsed, dict) else ({}, False)


def is_valid_output_json(record: dict[str, Any]) -> bool:
    return all(field in record for field in OUTPUT_FIELDS)


def normalize_record_fields(record: dict[str, Any]) -> dict[str, Any]:
    normalized: dict[str, Any] = {}
    for key, value in record.items():
        canonical_key = FIELD_ALIASES.get(key, key)
        normalized[canonical_key] = normalize_cell(value)
    return normalized


def first_present(record: dict[str, Any], keys: list[str]) -> Any:
    for key in keys:
        if key in record and record[key] not in (None, ""):
            return record[key]
    return None


def values_equal(predicted: Any, expected: Any) -> bool:
    predicted_normalized = normalize_for_compare(predicted)
    expected_normalized = normalize_for_compare(expected)
    if isinstance(predicted_normalized, list) and isinstance(expected_normalized, str):
        return predicted_normalized == [expected_normalized]
    if isinstance(predicted_normalized, str) and isinstance(expected_normalized, list):
        return [predicted_normalized] == expected_normalized
    return predicted_normalized == expected_normalized


def normalize_for_compare(value: Any) -> Any:
    value = normalize_cell(value)
    if value is None:
        return None
    if isinstance(value, list):
        items = sorted(
            normalize_text(str(item))
            for item in value
            if normalize_text(str(item)) and normalize_text(str(item)).lower() not in {"null", "none", "nan"}
        )
        return items or None
    if isinstance(value, (int, float)):
        return int(value) if float(value).is_integer() else value

    text = normalize_text(str(value))
    if not text:
        return None
    if "；" in text or ";" in text:
        parts = sorted(part for part in re.split(r"[；;]", text) if part)
        return parts or None
    if re.fullmatch(r"\d+(\.0+)?", text):
        return int(float(text))
    return text


def normalize_cell(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, str):
        text = value.strip()
        if text.lower() in {"", "null", "none", "nan"}:
            return None
        return text
    return value


def normalize_text(value: str) -> str:
    return re.sub(r"\s+", "", value.strip(" ，。；;：:、"))


def normalize_header(value: Any) -> str | None:
    if value is None:
        return None
    return str(value).strip()


def safe_rate(correct: int, total: int) -> float:
    if total == 0:
        return 0.0
    return round(correct / total, 4)


if __name__ == "__main__":
    main()
